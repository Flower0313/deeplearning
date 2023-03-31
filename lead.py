import numpy as np
import torch
import torch.nn as nn

import openpyxl
from torch.utils.data import Dataset

wb = openpyxl.load_workbook(r'T:\deeplearning\train\lead.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

result = np.empty((0, 4), dtype=int)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 跟进次数
    follow = sheet.cell(row=i, column=2)
    # 再联系
    contact = sheet.cell(row=i, column=3)
    # 已预约
    order = sheet.cell(row=i, column=4)
    # 已签单
    buy = sheet.cell(row=i, column=5)
    data = np.array([[follow.value, contact.value, order.value, buy.value]])
    result = np.concatenate([result, data], axis=0)


# 数据
class CrmDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = torch.from_numpy(data[:, :-1]).to(torch.float32)
        self.y_data = torch.from_numpy(data[:, [-1]]).to(torch.float32)

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


# 1. 定义神经网络模型
class Model(nn.Module):
    def __init__(self):
        super(Model, self).__init__()
        self.fc1 = nn.Linear(3, 32)
        self.fc2 = nn.Linear(32, 16)
        self.fc3 = nn.Linear(16, 1)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = self.fc3(x)
        return x


if __name__ == '__main__':
    my_dataset = CrmDataset(result)
    model = Model()
    # 损失
    criterion = torch.nn.MSELoss()
    optimizer = torch.optim.Adam(model.parameters())
    for epoch in range(10000):
        optimizer.zero_grad()
        y_pred = model(my_dataset.x_data)
        loss = criterion(y_pred, my_dataset.y_data)
        # 查看损失值
        print(epoch, loss.item())
        loss.backward()
        optimizer.step()

    x_test = torch.Tensor([[64986, 25862, 1053]])
    # 预测值
    y_test = model(x_test)
    print('y_pred', y_test.data)
