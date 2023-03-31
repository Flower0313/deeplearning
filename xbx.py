import math
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pymysql
import pandas as pd
import torch
from torch import nn
from torch.utils.data import Dataset

wb = openpyxl.load_workbook(r'T:\deeplearning\train\amount.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

result = np.empty((0, 4), dtype=int)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=1)
    status = sheet.cell(row=i, column=2)
    status2 = sheet.cell(row=i, column=3)
    status3 = sheet.cell(row=i, column=4)

    data = str(lesson.value).split(',')
    data.append(status.value)
    data.append(status2.value)
    data.append(status3.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)


class CrmDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = torch.from_numpy(data[:, :-1]).to(torch.float32)
        self.y_data = torch.from_numpy(data[:, [-1]]).to(torch.float32)

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


class Model(nn.Module):
    def __init__(self):
        super(Model, self).__init__()
        self.fc1 = nn.Linear(3, 16)
        self.fc2 = nn.Linear(16, 16)
        self.fc3 = nn.Linear(16, 1)

    def forward(self, x):
        x = self.fc1(x)
        x = self.fc2(x)
        x = self.fc3(x)
        return x


dataSet = CrmDataset(result)
model = Model()
# 损失
criterion = torch.nn.MSELoss()
# 优化
optimizer = torch.optim.Adam(model.parameters(), lr=0.1, weight_decay=1e-2)
losses = []
for epoch in range(1000):
    y_pred = model(dataSet.x_data)
    loss = criterion(y_pred, dataSet.y_data)
    losses.append(loss.item())
    optimizer.zero_grad()
    loss.backward()
    optimizer.step()

plt.cla()  # 清空
plt.scatter(dataSet.x_data[:, 2], dataSet.y_data, c='blue', s=50, alpha=0.3, label='train')

w = model.state_dict()['fc1.weight']
b = model.state_dict()['fc1.bias']
print(w, b)
# 画出预测模型
# with torch.no_grad():
#     x = []
#     for i in range(5, 30):
#         x.append([i, i * i, i * i * i])


plt.ioff()
plt.show()
print(losses[-1])
# plt.plot(list(range(1, len(losses) + 1)), losses, 'black', lw=3)
# plt.xlabel('number')
# plt.ylabel('loss')
# plt.show()
