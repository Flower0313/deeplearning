import math
import numpy as np
import openpyxl
import torch
import matplotlib.pyplot as plt
from torch import nn
from torch.utils.data import Dataset
import torch.nn.functional as F
from torch.utils.tensorboard import SummaryWriter

np.set_printoptions(suppress=True, precision=3)


class StuDataset(Dataset):
    def __init__(self, data):
        self.x_data = torch.from_numpy(normalize(data[:, :-1])).to(torch.float32)
        self.y_data = torch.from_numpy(data[:, [-1]]).to(torch.float32)

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


class Model(nn.Module):
    def __init__(self):
        super(Model, self).__init__()
        self.layer1 = nn.Linear(7, 16)
        self.layer2 = nn.Linear(16, 50)
        self.layer3 = nn.Linear(50, 1)

    def forward(self, x):
        x = self.layer1(x)
        x = self.layer2(x)
        x = self.layer3(x)
        # out = F.softmax(x, dim=1)
        return x


def normalize(array):
    # 计算均值
    means = np.mean(array, axis=0)
    # 计算标准差
    stds = np.std(array, axis=0)
    return (array - means) / stds


wb = openpyxl.load_workbook(r'T:\deeplearning\train\xkl.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
result = np.empty((0, 8), dtype=float)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=1)
    status = sheet.cell(row=i, column=2)
    status2 = sheet.cell(row=i, column=3)
    status3 = sheet.cell(row=i, column=4)
    status4 = sheet.cell(row=i, column=5)
    status5 = sheet.cell(row=i, column=6)
    status6 = sheet.cell(row=i, column=7)
    status7 = sheet.cell(row=i, column=8)

    data = str(lesson.value).split(',')
    data.append(status.value)
    data.append(status2.value)
    data.append(status3.value)
    data.append(status4.value)
    data.append(status5.value)
    data.append(status6.value)
    data.append(status7.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)

if __name__ == '__main__':
    print(result.shape)
    dataSet = StuDataset(result)
    model = Model()
    criterion = torch.nn.MSELoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=0.1, weight_decay=1e-2)
    for epoch in range(500):
        y_pred = model(dataSet.x_data)
        loss = criterion(y_pred, dataSet.y_data)
        print("loss", loss.item())
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()