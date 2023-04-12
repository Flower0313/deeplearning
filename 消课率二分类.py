import math
import numpy as np
import openpyxl
import torch
import matplotlib.pyplot as plt
from torch import nn
from torch.utils.data import Dataset


class BinaryClassificationModel(nn.Module):
    def __init__(self):
        super(BinaryClassificationModel, self).__init__()
        self.layer1 = nn.Linear(2, 1)
        # self.layer2 = nn.Linear(16, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        x = self.layer1(x)
        # x = self.layer2(x)
        x = self.sigmoid(x)
        return x


def decision_boundary(w, b, x1):
    print(w.numpy())
    x2 = (- w.numpy()[0][0] * x1) / w.numpy()[0][1]
    return x2


def normalize(array):
    # 计算均值
    means = np.mean(array, axis=0)
    # 计算标准差
    stds = np.std(array, axis=0)
    return (array - means) / stds


wb = openpyxl.load_workbook(r'T:\deeplearning\train\binary.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

result = np.empty((0, 3), dtype=int)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=5)
    status = sheet.cell(row=i, column=6)
    # status2 = sheet.cell(row=i, column=3)
    status3 = sheet.cell(row=i, column=7)

    data = str(lesson.value).split(',')
    data.append(status.value)
    # data.append(status2.value)
    data.append(status3.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)

if __name__ == '__main__':
    plt.scatter(x=normalize(result[:, 0].tolist()), y=normalize(result[:, 1].tolist()), marker='*',
                c=result[:, 2].tolist())

    model = BinaryClassificationModel()
    criterion = torch.nn.BCELoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=0.01)
    for epoch in range(500):
        # 前向传播
        outputs = model(torch.Tensor(result[:, 0:2]))
        loss = criterion(outputs,
                         torch.from_numpy(result[:, 2].reshape(len(result[:, 2].tolist()), 1)).to(torch.float32))
        # 反向传播和优化
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

    w = model.state_dict()['layer1.weight']
    b = model.state_dict()['layer1.bias']
    x1 = torch.linspace(-1, 5, 500)
    plt.xlabel('sales amount')
    plt.ylabel('arrange')
    plt.title("消课率80%上为优秀", fontname="SimHei")
    # torch.save(model.state_dict(), 'D:/trained_model/teacher_binary.pt')
    # print(x1)
    x2 = decision_boundary(w, b, x1)
    plt.plot(x1.tolist(), x2.tolist(), color="red")
    plt.show()
