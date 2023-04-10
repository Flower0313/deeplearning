import numpy as np
import openpyxl
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.metrics import accuracy_score
from sklearn.naive_bayes import MultinomialNB, ComplementNB, GaussianNB
from sklearn.model_selection import train_test_split

# 取消numpy科学计数法
np.set_printoptions(suppress=True)


class Dataset:
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = data[:9990, :-1]
        # self.y_data = data[:5350, [-1]]
        self.x_test = data[9990:, :-1]
        # self.y_test = data[5350:, [-1]]
        self.y_labels = np.reshape(data[:9990, [-1]], (1, 9990))[0]
        self.y_labels_test = np.reshape(data[9990:, [-1]], (1, 10))[0]


def createData():
    wb = openpyxl.load_workbook(r'T:\deeplearning\train\beyes2.xlsx')
    sheet = wb.get_sheet_by_name("Sheet1")
    result = np.empty((0, 12), dtype=float)
    # 遍历excel
    # 要使用python package,不然没有权限访问文件夹
    for i in range(2, sheet.max_row + 1):  #
        # 课堂
        lesson = sheet.cell(row=i, column=1)
        status = sheet.cell(row=i, column=2)
        status2 = sheet.cell(row=i, column=3)
        status3 = sheet.cell(row=i, column=4)
        status4 = sheet.cell(row=i, column=5)
        status5 = sheet.cell(row=i, column=6)
        status6 = sheet.cell(row=i, column=7)
        status7 = sheet.cell(row=i, column=8)
        status8 = sheet.cell(row=i, column=9)
        status9 = sheet.cell(row=i, column=10)
        status10 = sheet.cell(row=i, column=11)
        status11 = sheet.cell(row=i, column=12)

        data = str(lesson.value).split(',')
        data.append(status.value)
        data.append(status2.value)
        data.append(status3.value)
        data.append(status4.value)
        data.append(status5.value)
        data.append(status6.value)
        data.append(status7.value)
        data.append(status8.value)
        data.append(status9.value)
        data.append(status10.value)
        data.append(status11.value)
        data = np.array(data).reshape(1, len(data))
        result = np.concatenate([result, data], axis=0)
    return result


# print(dataSet.y_labels)
if __name__ == '__main__':
    dataSet = Dataset(createData())

    # 创建朴素贝叶斯分类器
    clf = GaussianNB()

    # 使用数据训练分类器
    clf.fit(dataSet.x_data, dataSet.y_labels)

    # 使用分类器进行预测
    y_pred = clf.predict(dataSet.x_test)

    # 打印出每种可能的概率
    proba = clf.predict_proba(dataSet.x_test)
    print("概率：", np.round(proba * 100, 3))
    print("推荐结果：", y_pred)
    print("实际结果：", dataSet.y_labels_test)
