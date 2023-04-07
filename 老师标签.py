import math
import numpy as np
import openpyxl
import torch
import matplotlib.pyplot as plt
from torch import nn
from torch.utils.data import Dataset
import torch.nn.functional as F
from torch.utils.tensorboard import SummaryWriter

np.set_printoptions(suppress=True, precision=3)


class StuDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = torch.from_numpy(data[:9800, :-1]).to(torch.float32)
        self.y_data = torch.from_numpy(data[:9800, [-1]]).to(torch.float32)
        self.x_test = torch.from_numpy(data[9800:, :-1]).to(torch.float32)
        self.y_test = torch.from_numpy(data[9800:, [-1]]).to(torch.float32)
        self.y_labels = torch.from_numpy(data[:9800, [-1]]).to(torch.float32).view(1, 9800)[0].long()
        self.y_labels_test = torch.from_numpy(data[9800:, [-1]]).to(torch.float32).view(1, 200)[0].long()

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


class BinaryClassificationModel(nn.Module):
    def __init__(self):
        super(BinaryClassificationModel, self).__init__()
        self.layer1 = nn.Linear(16, 50)
        self.layer2 = nn.Linear(50, 100)
        self.layer3 = nn.Linear(100, 50)
        self.layer4 = nn.Linear(50, 4)
        # self.softmax = nn.Softmax()

    def forward(self, x):
        x = self.layer1(x)
        x = self.layer2(x)
        x = self.layer3(x)
        x = self.layer4(x)
        # out = F.softmax(x, dim=1)
        return x


def normalize(array):
    # 计算均值
    means = np.mean(array, axis=0)
    # 计算标准差
    stds = np.std(array, axis=0)
    return (array - means) / stds


tb_writer = SummaryWriter(log_dir=r'T:\deeplearning\tensorboard\board')
wb = openpyxl.load_workbook(r'T:\deeplearning\train\multi.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
result = np.empty((0, 17), dtype=float)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=1)
    status = sheet.cell(row=i, column=2)
    status2 = sheet.cell(row=i, column=3)
    status3 = sheet.cell(row=i, column=4)
    status4 = sheet.cell(row=i, column=5)
    status5 = sheet.cell(row=i, column=6)
    status6 = sheet.cell(row=i, column=7)
    status7 = sheet.cell(row=i, column=8)
    status8 = sheet.cell(row=i, column=9)
    status9 = sheet.cell(row=i, column=10)
    status10 = sheet.cell(row=i, column=11)
    status11 = sheet.cell(row=i, column=12)
    status12 = sheet.cell(row=i, column=13)
    status13 = sheet.cell(row=i, column=14)
    status14 = sheet.cell(row=i, column=15)
    status15 = sheet.cell(row=i, column=16)
    status16 = sheet.cell(row=i, column=17)

    data = str(lesson.value).split(',')
    data.append(status.value)
    data.append(status2.value)
    data.append(status3.value)
    data.append(status4.value)
    data.append(status5.value)
    data.append(status6.value)
    data.append(status7.value)
    data.append(status8.value)
    data.append(status9.value)
    data.append(status10.value)
    data.append(status11.value)
    data.append(status12.value)
    data.append(status13.value)
    data.append(status14.value)
    data.append(status15.value)
    data.append(status16.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)

if __name__ == '__main__':
    model = BinaryClassificationModel()
    # 会自动进行softmax计算和交叉熵
    criterion = nn.CrossEntropyLoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=0.1, weight_decay=1e-2)
    result[:, 13] = normalize(result[:, 13])
    result[:, 14] = normalize(result[:, 14])
    result[:, 15] = normalize(result[:, 15])
    dataSet = StuDataset(result)
    losses = []
    for epoch in range(500):
        # 前向传播
        outputs = model(dataSet.x_data)
        # print(dataSet.y_data.view(1, 9800)[0].long())
        loss = criterion(outputs, dataSet.y_labels)
        # print(loss.item())
        losses.append(loss.item())
        tb_writer.add_scalar("train_loss", loss.item(), epoch)
        # 反向传播和优化
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
    tb_writer.add_graph(model, dataSet.x_data)

    # torch.save(model.state_dict(), 'D:/trained_model/teacher_multi.pt')
    with torch.no_grad():
        x_test = torch.Tensor(dataSet.x_test[:200, :])
        y_test = model(x_test)
        pred = y_test.argmax(dim=1)
        # print("预测值:", pred.detach().numpy())
        # print(dataSet.y_labels_test.detach().numpy())
        diff_count = np.count_nonzero(pred.detach().numpy() != dataSet.y_labels_test.detach().numpy())
        # print(diff_count)
        print((1 - diff_count / 200) * 100, "%")

    # plt.plot(list(range(1, len(losses) + 1)), losses, 'black', lw=3)
    # plt.xlabel('number')
    # plt.ylabel('loss')
    # plt.show()
