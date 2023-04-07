import operator

import openpyxl
import torch
import numpy as np
import pandas as pd

# 取消numpy科学计数法
from torch.utils.data import Dataset

np.set_printoptions(suppress=True)
# 取消tensor科学计数法
torch.set_printoptions(sci_mode=False)

# 再转为tensor之前先改变numpy中数据的类型
# stock_list = torch.tensor(stock_data.astype(float))
wb = openpyxl.load_workbook(r'T:\deeplearning\train\multi.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
result = np.empty((0, 17), dtype=float)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=1)
    status = sheet.cell(row=i, column=2)
    status2 = sheet.cell(row=i, column=3)
    status3 = sheet.cell(row=i, column=4)
    status4 = sheet.cell(row=i, column=5)
    status5 = sheet.cell(row=i, column=6)
    status6 = sheet.cell(row=i, column=7)
    status7 = sheet.cell(row=i, column=8)
    status8 = sheet.cell(row=i, column=9)
    status9 = sheet.cell(row=i, column=10)
    status10 = sheet.cell(row=i, column=11)
    status11 = sheet.cell(row=i, column=12)
    status12 = sheet.cell(row=i, column=13)
    status13 = sheet.cell(row=i, column=14)
    status14 = sheet.cell(row=i, column=15)
    status15 = sheet.cell(row=i, column=16)
    status16 = sheet.cell(row=i, column=17)

    data = str(lesson.value).split(',')
    data.append(status.value)
    data.append(status2.value)
    data.append(status3.value)
    data.append(status4.value)
    data.append(status5.value)
    data.append(status6.value)
    data.append(status7.value)
    data.append(status8.value)
    data.append(status9.value)
    data.append(status10.value)
    data.append(status11.value)
    data.append(status12.value)
    data.append(status13.value)
    data.append(status14.value)
    data.append(status15.value)
    data.append(status16.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)


class StuDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = data[:9800, :-1]
        self.y_data = data[:9800, [-1]]
        self.x_test = data[9800:, :-1]
        self.y_test = data[9800:, [-1]]

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


def normalize(array):
    # 计算均值
    means = np.mean(array, axis=0)
    # 计算标准差
    stds = np.std(array, axis=0)
    return (array - means) / stds


def classify(inX, dataSet, labels, k):
    dataSetSize = dataSet.shape[0]
    # 在列向量方向上重复inX共1次(横向),行向量方向上重复inX共dataSetSize次(纵向)
    diffMat = np.tile(inX, (dataSetSize, 1)) - dataSet
    # 二维特征相减后平方
    sqDiffMat = diffMat ** 2
    # sum()所有元素相加,sum(0)列相加,sum(1)行相加
    sqDistances = sqDiffMat.sum(axis=1)
    # 开方,计算出距离
    distances = sqDistances ** 0.5
    # 返回distances中元素从小到大排序后的索引值
    sortedDistIndices = distances.argsort()
    # 定一个记录类别次数的字典
    classCount = {}
    for i in range(k):
        # 取出前k个元素的类别
        voteIlabel = int(labels[sortedDistIndices[i]][0])
        # dict.get(key,default=None),字典的get()方法,返回指定键的值,如果值不在字典中返回默认值。
        # 计算类别次数
        classCount[voteIlabel] = classCount.get(voteIlabel, 0) + 1

    sortedClassCount = sorted(classCount.items(), key=operator.itemgetter(1), reverse=True)
    # 返回次数最多的类别,即所要分类的类别
    return sortedClassCount[0][0]


if __name__ == '__main__':
    result[:, 13] = normalize(result[:, 13])
    result[:, 14] = normalize(result[:, 14])
    result[:, 15] = normalize(result[:, 15])
    dataSet = StuDataset(result)
    j = 0
    error = 0
    for i in dataSet.x_test:
        classifierResult = classify(i, dataSet.x_data, dataSet.y_data, 20)
        real = int(dataSet.y_test[j, :][0])
        if real != classifierResult:
            error += 1
        print("预测结果:%s\t真实结果:%s" % (str(classifierResult), str(real)))
        j += 1

    print("正确率:", (1 - error / 200) * 100, "%")
