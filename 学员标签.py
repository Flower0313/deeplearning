import math
import numpy as np
import openpyxl
import torch
import matplotlib.pyplot as plt
from torch import nn
from torch.utils.data import Dataset


class StuDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = torch.from_numpy(data[:9800, :-1]).to(torch.float32)
        self.y_data = torch.from_numpy(data[:9800, [-1]]).to(torch.float32)
        self.x_test = torch.from_numpy(data[9800:, :-1]).to(torch.float32)
        self.y_test = torch.from_numpy(data[9800:, [-1]]).to(torch.float32)

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


class BinaryClassificationModel(nn.Module):
    def __init__(self):
        super(BinaryClassificationModel, self).__init__()
        self.layer1 = nn.Linear(6, 32)
        self.layer2 = nn.Linear(32, 32)
        self.layer3 = nn.Linear(32, 1)
        self.sigmoid = nn.Sigmoid()

    def forward(self, x):
        x = self.layer1(x)
        x = self.layer2(x)
        x = self.layer3(x)
        x = self.sigmoid(x)
        return x


def decision_boundary(w, b, x1):
    print(w.numpy())
    x2 = (- w.numpy()[0][0] * x1) / w.numpy()[0][1]
    return x2


def normalize(array):
    # 计算均值
    means = np.mean(array, axis=0)
    # 计算标准差
    stds = np.std(array, axis=0)
    return (array - means) / stds


wb = openpyxl.load_workbook(r'T:\deeplearning\train\student_binary.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")
result = np.empty((0, 7), dtype=float)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=1)
    status = sheet.cell(row=i, column=2)
    status2 = sheet.cell(row=i, column=3)
    status3 = sheet.cell(row=i, column=4)
    status4 = sheet.cell(row=i, column=5)
    status5 = sheet.cell(row=i, column=6)
    status6 = sheet.cell(row=i, column=7)

    data = str(lesson.value).split(',')
    print(type(data))
    data.append(status.value)
    data.append(status2.value)
    data.append(status3.value)
    data.append(status4.value)
    data.append(status5.value)
    data.append(status6.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)

if __name__ == '__main__':
    model = BinaryClassificationModel()
    criterion = torch.nn.BCELoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=0.1)
    result[:, 0] = normalize(result[:, 0])
    result[:, 1] = normalize(result[:, 1])
    result[:, 2] = normalize(result[:, 2])
    result[:, 3] = normalize(result[:, 3])
    dataSet = StuDataset(result)
    for epoch in range(400):
        # 前向传播
        outputs = model(dataSet.x_data)
        loss = criterion(outputs, dataSet.y_data)
        print(loss.item())
        # 反向传播和优化
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()

    plt.xlabel('amount')
    plt.ylabel('diff')
    plt.title("沉睡或活跃学员", fontname="SimHei")
    # torch.save(model.state_dict(), 'D:/trained_model/student_binary.pt')
    x_test = torch.Tensor(dataSet.x_test[:200, :])
    np.set_printoptions(suppress=True, precision=3)
    y_test = model(x_test)
    r = y_test.detach().numpy() * 100
    a = np.where(r[:, -1] >= 50, 1, 0)
    b = dataSet.y_test[:200, :].detach().numpy()[:, -1]
    diff_count = np.count_nonzero(a != b)
    my_list = list(range(1, 201))
    plt.plot(my_list, a, 'red', lw=1)
    plt.plot(my_list, b, 'blue', lw=1)
    plt.ioff()
    plt.show()
    print("正确率:", (1 - diff_count / 100) * 100, "%")
