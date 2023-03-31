import math

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pymysql
import pandas as pd
import torch
from torch import nn
from torch.utils.data import Dataset

# SELECT  *
# FROM    (
#             SELECT  teacher_id
#                     ,subject3_id
#                     ,city_code/10000
#                     ,TO_CHAR(lesson_start_time,'yyyy-mm') AS ym
#                     ,ROW_NUMBER() OVER (PARTITION BY teacher_id ORDER BY TO_CHAR(lesson_start_time,'yyyy-mm') ) AS px
#                     ,NVL(COUNT(DISTINCT CASE WHEN finish_status = 1 THEN student_id ELSE NULL END),0) AS stus
#             FROM    htdw.rp_lesson_finish_detail
#             WHERE   ds = ${bizdate}
#             GROUP BY teacher_id
#                      ,subject3_id
#                      ,city_code
#                      ,TO_CHAR(lesson_start_time,'yyyy-mm')
#             ORDER BY teacher_id,ym
#         ) a
# WHERE   stus <> 0
# ;

wb = openpyxl.load_workbook(r'T:\deeplearning\模型\teacher.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

result = np.empty((0, 4), dtype=int)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=1)

    status = sheet.cell(row=i, column=2)
    status2 = sheet.cell(row=i, column=3)
    status3 = sheet.cell(row=i, column=4)

    data = str(lesson.value).split(',')
    data.append(status.value)
    data.append(status2.value)
    data.append(status3.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)


class CrmDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = torch.from_numpy(data[:, :-1]).to(torch.float32)
        self.y_data = torch.from_numpy(data[:, [-1]]).to(torch.float32)

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


class Model(nn.Module):
    def __init__(self):
        super(Model, self).__init__()
        self.fc1 = nn.Linear(3, 32)
        self.fc2 = nn.Linear(32, 16)
        self.fc3 = nn.Linear(16, 1)

    def forward(self, x):
        x = self.fc1(x)
        x = self.fc2(x)
        x = self.fc3(x)
        return x


if __name__ == '__main__':
    dataSet = CrmDataset(result)
    model = Model()
    # 损失
    criterion = torch.nn.MSELoss()
    optimizer = torch.optim.Adam(model.parameters(), lr=0.01, weight_decay=1e-2)
    for epoch in range(500):
        y_pred = model(dataSet.x_data)
        loss = criterion(y_pred, dataSet.y_data)
        # print("loss", loss.item())
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
        if epoch % 100 == 0:
            plt.cla()  # 清空
            plt.scatter(dataSet.x_data[:, 2], dataSet.y_data, c='blue', s=50, alpha=0.3, label='train')
            plt.plot(dataSet.x_data[:, 2], y_pred.data.numpy(), 'red', lw=3)

    x_test = torch.Tensor([[8, 31.01, 13]])
    # 期望业绩
    y_test = model(x_test)
    print(y_test)
    plt.ioff()
    plt.show()
