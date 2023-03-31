import math

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pymysql
import pandas as pd
import torch
from torch import nn
from torch.utils.data import Dataset

# SELECT  lx,submenu_store_id,round((keyDataCount-min(keyDataCount) over())*10/(max(keyDataCount) over()-min(keyDataCount) over()),3) as value
# ,min(keyDataCount) over() as minmin
# ,max(keyDataCount) over() as minmin
# ,keyDataCount
# FROM    (
#             SELECT  *
#                     ,COUNT(*) OVER (PARTITION BY staff_id ) AS nums
#                     ,MAX(lx) OVER (PARTITION BY staff_id ) AS maxn
#                     ,min(keyDataCount) over (partition by staff_id) as minprice
#             FROM    (
#                         SELECT  *
#                         FROM    (
#                                     SELECT  *
#                                             ,SUM(last_date) OVER (PARTITION BY staff_id ORDER BY zr_month ) AS lx
#                                     FROM    (
#                                                 SELECT  *
#                                                         ,NVL(DATEDIFF(CONCAT(zr_month,'-01 00:00:00'),CONCAT(LAG(zr_month,1,NULL) OVER (PARTITION BY staff_id ORDER BY zr_month ),'-01 00:00:00'),'mm'),1) AS last_date
#                                                 FROM    (
#                                                             SELECT  a.staff_id
#                                                                     ,a.submenu_store_id
#                                                                     ,a.zr_month
#                                                                     ,NVL(SUM(CASE WHEN operation_type <> -1 AND staff_id >= 0 THEN bcs_score ELSE 0 END),0) - NVL(SUM(CASE WHEN operation_type = -1 AND staff_id >= -1 THEN bcs_score ELSE 0 END),0) AS keyDataCount
#                                                             FROM    htdw.rp_dw_contract_sales a
#                                                             left join htdw.ods_by_staff b
#                                                             on a.staff_id=b.id
#                                                             and b.ds=${bizdate}
#                                                             WHERE   a.ds = ${bizdate}
#                                                             AND     a.staff_id > 0
#                                                             GROUP BY a.staff_id
#                                                                      ,a.submenu_store_id
#                                                                      ,a.zr_month
#                                                             ORDER BY a.staff_id,zr_month
#                                                         ) a
#                                             ) b
#                                 ) c
#                         WHERE   lx <= 6
#                     ) d
#         ) f
# WHERE   nums = 6
# AND     maxn = 6
# and minprice>=0
# ;


wb = openpyxl.load_workbook(r'T:\deeplearning\train\cc.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

result = np.empty((0, 3), dtype=int)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=1)

    status = sheet.cell(row=i, column=2)
    status2 = sheet.cell(row=i, column=3)
    # status3 = sheet.cell(row=i, column=3)

    data = str(lesson.value).split(',')
    data.append(status.value)
    data.append(status2.value)
    # data.append(status3.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)


class CrmDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = torch.from_numpy(data[:, :-1]).to(torch.float32)
        self.y_data = torch.from_numpy(data[:, [-1]]).to(torch.float32)

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


class Model(nn.Module):
    def __init__(self):
        super(Model, self).__init__()
        self.fc1 = nn.Linear(2, 16)
        self.fc2 = nn.Linear(16, 16)
        self.fc3 = nn.Linear(16, 1)

    def forward(self, x):
        x = self.fc1(x)
        x = self.fc2(x)
        x = self.fc3(x)
        return x


dataSet = CrmDataset(result)
model = Model()
# 损失
criterion = torch.nn.MSELoss()
optimizer = torch.optim.Adam(model.parameters(), lr=0.1, weight_decay=1e-2)
for epoch in range(1000):
    y_pred = model(dataSet.x_data)
    loss = criterion(y_pred, dataSet.y_data)
    # print("loss", loss.item())
    optimizer.zero_grad()
    loss.backward()
    optimizer.step()

plt.cla()  # 清空
plt.scatter(dataSet.x_data[:, 0], dataSet.y_data, c='blue', s=50, alpha=0.3, label='train')
plt.title("cc入职前半年理想完成曲线",fontname = "SimHei")
plt.xlabel("month")
plt.ylabel("sales quota")
# plt.plot(dataSet.x_data[:, 0], y_pred.data.numpy(), 'red', lw=3)

# 画出预测模型
with torch.no_grad():
    x = []
    for i in range(1, 7):
        x.append([i, i * i])
    x_test = torch.Tensor(x)
    y_test = model(x_test)
    # 期望业绩
    plt.plot(x_test[:, 0], y_test.detach().numpy(), 'red', lw=3)

    x_test = torch.Tensor([[1, 1]])
    # 期望业绩
    y_test = model(x_test)
    print((y_test.detach().numpy() / 10 * 755623) - 28890)
# torch.save(model.state_dict(), 'D:/trained_model/cc.pt')
plt.ioff()
plt.show()
