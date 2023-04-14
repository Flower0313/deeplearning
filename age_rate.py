import math

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pymysql
import pandas as pd
import torch
from torch import nn
from torch.utils.data import Dataset

wb = openpyxl.load_workbook(r'T:\deeplearning\train\age_rate.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

result = np.empty((0, 4), dtype=int)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    # 课堂
    lesson = sheet.cell(row=i, column=1)
    status = sheet.cell(row=i, column=2)
    status2 = sheet.cell(row=i, column=3)
    status3 = sheet.cell(row=i, column=4)
    #status4 = sheet.cell(row=i, column=5)

    data = str(lesson.value).split(',')
    data.append(status.value)
    data.append(status2.value)
    data.append(status3.value)
    #data.append(status4.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)


class CrmDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.x_data = torch.from_numpy(normalize(data[:, :-1])).to(torch.float32)
        self.y_data = torch.from_numpy(data[:, [-1]]).to(torch.float32)

    def __getitem__(self, index):
        return self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


def normalize(array):
    # 计算均值
    means = np.mean(array, axis=0)
    # 计算标准差
    stds = np.std(array, axis=0)
    return (array - means) / stds


class Model(nn.Module):
    def __init__(self):
        super(Model, self).__init__()
        self.fc1 = nn.Linear(3, 16)
        self.fc3 = nn.Linear(16, 1)

    def forward(self, x):
        x = self.fc1(x)
        # x = self.fc2(x)
        x = self.fc3(x)
        return x


dataSet = CrmDataset(result)

model = Model()
# 损失
criterion = torch.nn.MSELoss()
optimizer = torch.optim.Adam(model.parameters(), lr=0.1, weight_decay=1e-2)
for epoch in range(5000):
    y_pred = model(dataSet.x_data)
    loss = criterion(y_pred, dataSet.y_data)
    # print("loss", loss.item())
    optimizer.zero_grad()
    loss.backward()
    optimizer.step()

plt.cla()  # 清空
plt.scatter(dataSet.x_data[:, 0], dataSet.y_data, c='blue', s=50, alpha=0.3, label='train')
plt.title("司龄和消课率", fontname="SimHei")
plt.xlabel("司龄", fontname="SimHei")
plt.ylabel("消课率", fontname="SimHei")
# plt.plot(dataSet.x_data[:, 0], y_pred.data.numpy(), 'red', lw=3)


# 打印生成的随机数
with torch.no_grad():
    x = []
    last_value = -1.2
    for i in range(1, 100):
        now_value = last_value + 0.04
        x.append([now_value, now_value * now_value, now_value ** 3])
        last_value = now_value
    x_test = torch.Tensor(x)
    y_test = model(x_test)
    # 期望业绩
    plt.plot(x_test[:, 0], y_test.detach().numpy(), 'red', lw=3)
plt.ioff()
plt.show()
