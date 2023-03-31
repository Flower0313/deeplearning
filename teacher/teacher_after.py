import numpy as np
import pymysql
import torch
import torch.nn as nn

import openpyxl
from torch.utils.data import Dataset

base_sql = '''
                insert into htdw_bi.rp_forecast_all values("{}","{}","{}","{}","{}")
            '''

wb = openpyxl.load_workbook(r'../train/teacher_test.xlsx')
sheet = wb.get_sheet_by_name("Sheet1")

result = np.empty((0, 10), dtype=float)
# 遍历excel
# 要使用python package,不然没有权限访问文件夹
for i in range(2, sheet.max_row + 1):  #
    teacher = sheet.cell(row=i, column=1)
    # 课堂
    lesson = sheet.cell(row=i, column=2)

    status = sheet.cell(row=i, column=4)

    data = str(lesson.value).split(',')
    data.append(status.value)
    data.insert(0, teacher.value)
    data = np.array(data).reshape(1, len(data)).astype(np.float)
    result = np.concatenate([result, data], axis=0)


# 数据
class CrmDataset(Dataset):
    def __init__(self, data):
        self.len = data.shape[0]
        self.teacher = torch.from_numpy(data[:, [0]])
        self.x_data = torch.from_numpy(data[:, 1:-1]).to(torch.float32)
        self.y_data = torch.from_numpy(data[:, [-1]]).to(torch.float32)

    def __getitem__(self, index):
        return self.teacher[index], self.x_data[index], self.y_data[index]

    def __len__(self):
        return self.len


# 1. 定义神经网络模型
class Model(nn.Module):
    def __init__(self):
        super(Model, self).__init__()
        self.fc1 = nn.Linear(8, 32)
        self.fc2 = nn.Linear(32, 16)
        self.fc3 = nn.Linear(16, 1)

    def forward(self, x):
        x = torch.relu(self.fc1(x))
        x = torch.relu(self.fc2(x))
        x = self.fc3(x)
        return x


if __name__ == '__main__':
    conn = pymysql.connect(
        host='rm-bp130tpgwlw1hkbbn8o.mysql.rds.aliyuncs.com',
        port=3306,
        user='user_bi_wr',
        passwd='BykJ_0824!',
        charset='utf8'
    )
    cursor = conn.cursor()
    new_data = CrmDataset(result)

    model = Model()
    # 加载训练好的模型
    model.load_state_dict(torch.load('D:/trained_model/teacher_model.pt'))
    # 将模式设置为预测模式
    model.eval()
    # 预测模型确保不要计算梯度
    with torch.no_grad():
        # 加载预测数据
        for x, y, z in new_data:
            outputs = model(y)
            sql = base_sql.format(str(x.item()).split(".")[0], -1, outputs.item(), 3, "老师离职率")
            cursor.execute(sql)
    # 重新训练模型,设置为训练模式
    # model.train()
    conn.commit()
    cursor.close()
    conn.close()

    # y_test = model(x_test)
    # if 1 <= y_test <= 2:
    #     print('y_pred', y_test.data)
    #     print('离职率', round((y_test.data - 1) * 100, 2) + "%")
    # elif y_test > 2:
    #     print('离职率', '接近100%')
    # else:
    #     print('离职率', '接近0%')
